#####-----------------------------------
#This scription can find the intersection of two cited aritles

#####-----------------------------------
import openpyxl as op
import xlrd
def GetTitle(filename,Sheetname):
    # wb=op.load_workbook(filename)
    # sh=wb["savedrecs"]
    # i = sh.cell(row=2,column=0).value
    # print(i)
    data = xlrd.open_workbook(filename)
    table = data.sheet_by_name(Sheetname)
    nrows = table.nrows
    ncols = table.ncols

    titleList = table.col_values(9,start_rowx=1,end_rowx=None)#Article1 Title column
    return titleList
#------------article 1 title
##please input articles' filename and sheetname
filename=r'C:\Users\LKChen\Desktop\savedrecs.xls'
Sheetname="savedrecs"
article1_list=GetTitle(filename,Sheetname)
filename2=r'C:\Users\LKChen\Desktop\savedrecs2.xls'
Sheetname2="savedrecs"
article2_list=GetTitle(filename2,Sheetname2)

#find intersection
list = set(article1_list).intersection(set(article2_list))


# print(list)
# print(len(list))

#---------output 
filename3=r'C:\Users\LKChen\Desktop\com-cite_list.xlsx'
wb=op.load_workbook(filename3)
sh=wb["Sheet1"]
j=2
sh.cell(row=1,column=1).value=str("Title")
for title in list:   
    sh.cell(row=j,column=1).value = title
    j=j+1
wb.save(filename3)